#!/usr/bin/env python3
"""Consolidate the pooled study's result CSVs into one .xlsx for Drive/Sheets.

Every sheet is derived from a file under journal_plan/, so this is a VIEW, not a
second source of truth — re-run it after any measurement lands and the workbook
catches up. Nothing is hand-typed.

Sheets
  00_README           what each sheet is, and the caveats that must travel with it
  01_Table1_gFID      generation FID, paper layout (Model x modality + Avg.)
  02_Table2_rFID      reconstruction FID, same layout
  03_gFID_raw         the full 48-row grid, all four extractors + null floor
  04_guidance         the g sweep that fixed g* = 3.0
  05_rFID_percell     the three per-cell files, melted to long format
  06_recon_slices     pooled-slice recon, orig vs DFT decoder (empty until measured)
  07_latent_geometry  per-modality latent statistics

WHY THE README SHEET EXISTS: two numbers in this book cannot be compared even
though they sit in the same layout. Table 1 is a single FID over one pooled mixed
sample; Table 2 is a mean of per-cell FIDs at each arm's own checkpoint. A reader
who divides one by the other gets a meaningless ratio, so the caveat ships inside
the file rather than in a chat message.

Usage:  python3 scripts/build_results_workbook.py [--out PATH]
"""
from __future__ import annotations

import argparse
import os

import pandas as pd
from openpyxl.styles import Alignment

JP = "journal_plan"
ARMS = [("maisi", "MAISI"), ("sid-cor50", "Ours (SID)"), ("vad-cov1var1", "Ours (VAD)")]
MODS = ["all_T1", "all_T2", "all_FLAIR"]
MOD_LABEL = {"all_T1": "All T1", "all_T2": "All T2", "all_FLAIR": "All FLAIR"}
# T1c is absent: brats_T1c is vae_only, outside the diffusion modality vocabulary.
CELLS_BY_MOD = {
    "all_T1": ["ukb_T1", "adni_T1", "oasis_T1", "brats_T1", "hcp_T1", "ixi_T1"],
    "all_T2": ["brats_T2", "hcp_T2", "ixi_T2", "oasis_T2"],
    "all_FLAIR": ["ukb_FLAIR", "adni_FLAIR", "brats_FLAIR"],
}
PERCELL = {
    "maisi": ("results_maisi_240_percell.csv", "checkpoint-240000"),
    "sid-cor50": ("results_sid_cor50_280_percell.csv", "checkpoint-280000"),
    "vad-cov1var1": ("results_vad_cov1var1_320_percell.csv", "checkpoint-320000"),
}

README = [
    ("Sheet", "What it is", "Read it with this caveat"),
    ("00_recon_percell",
     "Per-cell reconstruction, one row per eval run (row format, per "
     "results_datamix_percell_s1.csv). All FID columns are center ratio 0.4; "
     "ratio 1.0 is excluded from this book entirely. CELL granularity only.",
     "Two measurement sets, kept as SEPARATE rows because their checkpoints differ: "
     "LPIPS/PSNR/SSIM at 320k for every arm, and ratio-0.4 FID at each arm's own "
     "point (maisi 240k / sid 280k / vad 320k). Only VAD has ratio-0.4 FID at 320k, "
     "the checkpoint the diffusion latents came from - that gap is the checkpoint "
     "mismatch behind Table 2. The per-PLANE ratio-0.4 columns are blank: those "
     "numbers were measured and sit in the GSDS logs (compute_metric.py logs "
     "FID XY/YZ/ZX) but only the Avg was transcribed. Recovering the planes is a "
     "log parse. "
     "SLICE-level results (all_T1/T2/FLAIR) do NOT belong here - all_T1 contains "
     "ukb_T1, so one `cell` column holding both invites double counting. They go to "
     "07_recon_slices, which carries the same decoder axis."),
    ("05_Table1_gFID",
     "Generation FID. Inception, center ratio 0.4, guidance 3.0, n=2500 per cell. "
     "Real reference = the pre-shuffled TRAIN slice.",
     "One FID over one pooled mixed sample. Bold = best in column."),
    ("06_Table2_rFID",
     "Reconstruction FID, same extractor and ratio.",
     "A MEAN OF PER-CELL FIDs, and each arm sits at its own checkpoint "
     "(MAISI 240k / SID 280k / VAD 320k). NOT comparable to Table 1 in absolute "
     "value - different quantity, different checkpoint. Only the ordering and "
     "where the ordering changes can be read across the two."),
    ("07_gFID_raw", "The full 48-row grid: 3 modality slices + 13 cohort-modality cells.",
     "Cell rows are NOT interpretable: the model receives no cohort condition, so a "
     "cell whose metadata overlaps the bulk (oasis_T1: 95.8%) cannot be produced even "
     "by a perfect cohort-blind generator, and the null floor beside it is unreachable. "
     "Modality rows do not have this problem."),
    ("08_guidance", "Guidance sweep on all_T1, n=500, that fixed g* = 3.0.",
     "SID's curve is still descending at the swept edge (g=3.0), so SID's main-grid "
     "numbers were recorded away from its own optimum."),
    ("09_rFID_percell", "The per-cell reconstruction tables, melted to long format.",
     "Checkpoints differ per arm - see the ckpt column before comparing arms."),
    ("01_recon_perslice", "Pooled-slice reconstruction, original 320k decoder vs DFT 40k.",
     "Same schema as 00_recon_percell with `cell` -> `slice`. Empty until "
     "scripts/eval_recon_slices.sh runs on GSDS. This is what makes "
     "Table 1 and Table 2 finally comparable, and it is the only test of whether "
     "DFT helped - the DFT run logged no validation metric at all."),
    ("10_latent_by_modality", "Per-modality latent statistics from the saved embeddings.",
     "Measured on the SAMPLED latent z, not z_mu; the aux losses act on z_mu. "
     "Independent channel noise inflates the diagonal only, so off-diagonal "
     "correlation reads ~1% low for MAISI and ~0.25% low for SID/VAD."),
    ("SSIM warning",
     "Plain `ssim` is not comparable across these arms. VAD reads 0.699 while MAISI "
     "and SID read 0.985/0.984 - yet VAD wins PSNR (34.04 vs 32.61/32.98), LPIPS "
     "(0.021 vs 0.024/0.028) and SSIM_FG (0.952, the highest of the three).",
     "SSIM over the whole volume is dominated by the mostly-empty background, so a "
     "decoder whose background differs slightly collapses it while every foreground "
     "metric stays excellent. MAISI and SID have ssim > ssim_fg (background helps); "
     "VAD alone is inverted. Use `ssim_fg`, and never rank arms on plain `ssim`."),
    ("04_latent_stats",
     "One row per arm and the ONLY place its config and latent statistics live: "
     "kl/lambda, scale_factor + global_mean/std + avg_single_sigma "
     "(results_pooled_stats.csv), the channel-geometry panel "
     "(results_pooled_geometry.csv), and the absolute path of every checkpoint.",
     "Start here when a number elsewhere is ambiguous. Paths are the provenance: "
     "two confusions this session traced back to not recording which checkpoint a "
     "row came from."),
    ("", "", ""),
    ("Extractor note",
     "Inception r=0.4 is the only extractor validated against fidelity "
     "(rho=+0.70 over 66 points). SwAV -0.05, DINOv2 -0.00, RadImageNet -0.19.",
     "RadImageNet at ratio 1.0 is quality-blind (19.28 vs 19.29 on models whose SSIM "
     "was 0.315 vs 0.966) and is never quoted. The other extractors ship in "
     "03_gFID_raw as a panel and must not drive the ranking - they disagree: "
     "Inception says VAD, SwAV/DINOv2 say MAISI, RadImageNet says SID."),
]


STAGE1 = "/data/wonyoungjang/decodata/pooled/stage1"
# Identity, the config that produced it, and where every artefact lives. Borrowed
# from the MIUA book's "Weight Path" column, which was the one thing in it that
# made a row self-documenting. This session lost time twice for want of it: Table
# 2 turned out to mix three checkpoints, and the trajectory's FID_inc turned out
# to be the ukb_T1 cell rather than an average.
ARM_SPEC = [
    # key, label, lambda_cov, lambda_cor, lambda_var
    ("maisi",        "MAISI",      0.0, 0.0,  0.0),
    ("sid-cor50",    "Ours (SID)", 0.0, 50.0, 0.0),
    ("vad-cov1var1", "Ours (VAD)", 1.0, 0.0,  1.0),
]
KL_WEIGHT = 8e-4          # shared by all three arms (vae_train_stage1_eff32_kl8e4.json)
STAGE1_STEP = 320000      # the checkpoint the diffusion latents were extracted from


# Row schema copied from results_datamix_percell_s1.csv: one row per eval run,
# metrics as COLUMNS. The MIUA book put metrics in rows and models in columns,
# which reads well as a paper table but cannot be sorted, filtered or charted, and
# grew to 90 row-labels of copy-pasted 4-row blocks. Long format scales.
#
# NAMING follows that file: `rad_*` is center ratio 0.4. Ratio 1.0 columns are
# deliberately ABSENT — that setting is quality-blind and carrying it "as data with
# a caveat" just means it eventually gets read as a result.
#
# WHAT IS EMPTY AND WHY. compute_metric.py does emit per-plane FID (it logs
# "FID XY/YZ/ZX" at compute_metric.py:1309-1312), but the r=0.4 extractor-swap
# study transcribed only the aggregate into results_*_percell.csv. So rad_fid_* and
# inc_fid_* per-plane columns are blank here: the numbers were measured and live in
# the GSDS run logs, they were simply never carried into a CSV. Recovering them is
# a log parse, NOT a re-measurement.
RECON_COLS = [
    "exp", "ckpt", "decoder", "eval_split", "cohort", "modality", "cell", "n_images",
    "lpips", "psnr", "ssim", "ssim_fg",
    "rad_fid_xy", "rad_fid_yz", "rad_fid_zx", "rad_fid_avg",
    "inc_fid_xy", "inc_fid_yz", "inc_fid_zx", "inc_fid_avg",
    "swav_fid", "dino_fid_avg",
    "source",
]


# Hand-filled results promoted to a source. The per-plane ratio-0.4 FIDs were
# recovered from the GSDS run logs by hand on 2026-09-03; regenerating this sheet
# from the raw CSVs would silently destroy them. So: if the promoted file exists it
# IS the sheet, and the derivation below runs only to bootstrap it.
# EDIT THIS FILE, NOT pooled_results_csv/00_recon_percell.csv — the latter is an
# export and is rewritten on every run.
RECON_SOURCE = os.path.join(JP, "00_results_pooled_recon_percell.csv")
SLICE_SOURCE   = os.path.join(JP, "01_results_pooled_recon_perslice.csv")
GEN_CELL_SOURCE  = os.path.join(JP, "02_results_pooled_gen_percell.csv")
GEN_SLICE_SOURCE = os.path.join(JP, "03_results_pooled_gen_perslice.csv")
# Derived, not hand-maintained: every value comes from results_pooled_stats.csv,
# results_pooled_geometry.csv, or the ARM_SPEC literal above. Written out so the
# numbered set in journal_plan/ is complete, but regenerating it is always safe.
ARMS_OUT = os.path.join(JP, "04_latent_stats.csv")
def recon_percell_sheet() -> pd.DataFrame:
    if os.path.exists(RECON_SOURCE):
        return pd.read_csv(RECON_SOURCE)
    return _derive_recon_percell()


def _derive_recon_percell() -> pd.DataFrame:
    # Config and latent statistics are NOT repeated here — they live once in 01_arms,
    # joined by exp + ckpt. Copying 3 facts onto 84 rows is how the MIUA book's two
    # metric sheets drifted apart.
    def blank(key, ckpt, cell, n, split, src):
        cohort, _, mod = cell.partition("_")
        return {**{c: None for c in RECON_COLS},
                "exp": f"pooled-{key}-kl8e4-eff32-s1", "ckpt": ckpt,
                # Every row here decoded through the ORIGINAL stage1 decoder. That was
                # an unwritten assumption until DFT produced a second decoder per arm;
                # naming it makes a future per-cell DFT run just more rows, and nobody
                # has to guess which decoder an old row used.
                "decoder": "orig", "eval_split": split, "cohort": cohort, "modality": mod,
                "cell": cell, "n_images": n, "source": src}

    rows = []
    # (a) ratio-1.0 per-plane + LPIPS/PSNR/SSIM, all arms at the SAME checkpoint
    #     (320k = where the diffusion latents came from), split=test.
    cells_path = os.path.join(JP, "results_pooled_recon_s1_cells.csv")
    if os.path.exists(cells_path):
        d = pd.read_csv(cells_path)
        for key, *_ in ARM_SPEC:
            sub = d[d.exp.eq(f"pooled-{key}-kl8e4-eff32-s1-test-ck320000")]
            for _, r in sub.iterrows():
                row = blank(key, "checkpoint-320000", r["cell"], int(r["n_images"]),
                            "test", "results_pooled_recon_s1_cells.csv")
                # Only LPIPS/PSNR/SSIM survive from this source: its FID columns are
                # RadImageNet at center ratio 1.0, which is quality-blind (19.28 vs
                # 19.29 on models whose SSIM was 0.315 vs 0.966) and is never quoted.
                # The rows stay because these three ARE checkpoint-matched at 320k
                # across all arms, which nothing else in the book is.
                row.update(lpips=r["lpips"], psnr=r["psnr"], ssim=r["ssim"])
                rows.append(row)
    # (b) ratio-0.4 aggregates from the extractor-swap study. NOTE the checkpoints
    #     differ per arm — that is why these are separate rows, not extra columns on
    #     the rows above: merging them would imply one run produced both.
    for key, (fname, ckpt) in PERCELL.items():
        path = os.path.join(JP, fname)
        if not os.path.exists(path):
            continue
        w = pd.read_csv(path, index_col=0)
        for cell in w.columns:
            row = blank(key, ckpt, cell, None, "test", fname)
            row.update(lpips=w.loc["LPIPS", cell], psnr=w.loc["PSNR", cell],
                       ssim=w.loc["SSIM", cell], ssim_fg=w.loc["SSIM_FG", cell],
                       # The single number each extractor recorded IS the plane
                       # average: compute_metric.py:1315 logs
                       # "FID Avg: (fid_xy + fid_yz + fid_zx)/3", and *_fid_avg in
                       # results_datamix_percell_s1.csv equals mean(xy,yz,zx) to 1e-15
                       # over all 96 rows. So it belongs in _avg, not a parallel column.
                       # SwAV has no planes by construction — the Woodland protocol is
                       # axial-only, so avg == xy (compute_metric.py:994).
                       rad_fid_avg=w.loc["RadImgNet", cell], inc_fid_avg=w.loc["ImgNet", cell],
                       swav_fid=w.loc["SwAV", cell], dino_fid_avg=w.loc["DINOv2", cell])
            rows.append(row)
    df = pd.DataFrame(rows)[RECON_COLS]
    # One row per (exp, ckpt, decoder, cell). Two independent runs land on the same
    # key only for VAD, which both sources measured at 320k; their numbers differ
    # slightly (lpips 0.0214 vs 0.0210) because they are separate draws. Keep the
    # extractor-swap row: it is a strict superset, carrying ssim_fg and the four
    # ratio-0.4 FIDs on top of the same lpips/psnr/ssim.
    df["_rank"] = (df["source"] == "results_pooled_recon_s1_cells.csv").astype(int)
    df = (df.sort_values(["exp", "ckpt", "cell", "_rank"], kind="stable")
            .drop_duplicates(["exp", "ckpt", "decoder", "cell"], keep="first")
            .drop(columns="_rank"))
    return df.sort_values(["exp", "ckpt", "cell"], kind="stable").reset_index(drop=True)


def arms_sheet() -> pd.DataFrame:
    # results_pooled_stats.csv is the authoritative source for scale_factor and
    # friends — the per-run analysis/latent_stats.csv agrees with it exactly (checked
    # to 1e-9 for all three arms), and the W&B log values are trend-only.
    stats = pd.read_csv(os.path.join(JP, "results_pooled_stats.csv")).set_index("run_name")
    geom = pd.read_csv(os.path.join(JP, "results_pooled_geometry.csv")).set_index("run_name")
    GEO_COLS = ["raw_cos_sim", "diff_cos_sim", "offdiag_cov", "mean_l2_norm",
                "effective_rank", "avg_pairwise_cos_sim"]
    rows = []
    for key, label, lcov, lcor, lvar in ARM_SPEC:
        arm = f"pooled-{key}-kl8e4-eff32-s1"
        st = stats.loc[arm] if arm in stats.index else None
        ge = geom.loc[arm] if arm in geom.index else None
        sf = float(st["scaling_factor"]) if st is not None else None
        sig = float(st["avg_single_sigma"]) if st is not None else None
        sd = float(st["global_std"]) if st is not None else None
        rows.append({
            # exp + ckpt are the join key back to 00_recon_percell. The statistics
            # below are properties of the latents, and those were extracted from
            # STAGE1_STEP — so the ckpt is not decoration, it is what they describe.
            "exp": arm, "ckpt": f"checkpoint-{STAGE1_STEP}",
            "kl_weight": KL_WEIGHT,
            "lambda_cov": lcov, "lambda_cor": lcor, "lambda_var": lvar,
            "scaling_factor": round(sf, 4) if sf else None,
            "global_mean": round(float(st["global_mean"]), 4) if st is not None else None,
            "global_std": round(sd, 4) if sd else None,
            "avg_single_sigma": round(sig, 4) if sig else None,
            **{c: (round(float(ge[c]), 4) if ge is not None else None) for c in GEO_COLS},
            # sigma/std, not raw sigma: scale_factor normalises the latent to unit
            # variance before the diffusion sees it, so the relative posterior noise
            # is the quantity that survives. Raw sigma ranks SID and VAD apart
            # (0.059 vs 0.050) when normalised they are nearly tied (4.7% vs 5.1%).
            "sigma_over_std_pct": round(100 * sig / sd, 2) if (sig and sd) else None,
            "vae_ckpt": f"{STAGE1}/{arm}/weights/vae/checkpoint-{STAGE1_STEP}/model.pt",
            "dft_ckpt": f"{STAGE1}/{arm}-decft-n1.0/weights/vae/checkpoint-40000/model.pt",
            "unet_B_ckpt": f"{STAGE1}/{arm}/weights/unet/checkpoint-250000/model.pt",
            "unet_A_ckpt": f"{STAGE1}/{arm}-Acfg/weights/unet/checkpoint-250000/model.pt",
            "stats_source": os.path.join(JP, "results_pooled_stats.csv"),
            "geometry_source": os.path.join(JP, "results_pooled_geometry.csv"),
        })
    return pd.DataFrame(rows)


def paper_table(values: dict) -> pd.DataFrame:
    """values[(arm_key, slice)] -> float, laid out as Model x modality + Avg."""
    rows = []
    for key, label in ARMS:
        r = {"Model": label}
        vals = []
        for m in MODS:
            v = values.get((key, m))
            r[MOD_LABEL[m]] = round(v, 2) if v is not None else None
            if v is not None:
                vals.append(v)
        r["Avg."] = round(sum(vals) / len(vals), 2) if len(vals) == len(MODS) else None
        rows.append(r)
    return pd.DataFrame(rows)[["Model"] + [MOD_LABEL[m] for m in MODS] + ["Avg."]]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(JP, "pooled_results.xlsx"))
    # xlsx does not render in every viewer, and a CSV per sheet is also what you
    # want for diffing and for pasting a single sheet into a doc.
    # The numbered 00-04 set on its own: the files that are actually filled in by
    # hand (00-03) plus the arm identities they key against (04). The full workbook
    # additionally carries derived views and raw upstream CSVs, which is more than you
    # want open while entering measurements.
    ap.add_argument("--core-out", default=os.path.join(JP, "pooled_results_core.xlsx"),
                    help="also write a workbook of just the numbered 00-04 sheets")
    ap.add_argument("--csv-dir", default=os.path.join(JP, "pooled_results_csv"),
                    help="also write one CSV per sheet here (set empty to skip)")
    args = ap.parse_args()

    sheets: dict[str, pd.DataFrame] = {}
    sheets["99_README"] = pd.DataFrame(README[1:], columns=list(README[0]))
    sheets["00_recon_percell"] = recon_percell_sheet()

    # --- 01 Table 1: generation -------------------------------------------
    g = pd.read_csv(os.path.join(JP, "results_pooled_gfid.csv"))
    gmap = {(r["arm"], r["slice"]): float(r["fid_inception"]) for _, r in g.iterrows()}
    t1 = paper_table(gmap)
    # A reference row, the way the MIUA book carried a "Natural" row: the FID two
    # disjoint halves of the REAL data score at this n. It is what a perfect
    # generator would get, so it turns every number above it from an unanchored
    # magnitude into a multiple of the floor (T1 49x, T2 45x, FLAIR 66x at best).
    nulls = {m: float(g[g["slice"] == m]["fid_null_inception"].iloc[0]) for m in MODS}
    t1.loc[len(t1)] = ["null floor (real vs real)"] + \
        [round(nulls[m], 3) for m in MODS] + [round(sum(nulls.values()) / 3, 3)]
    sheets["05_Table1_gFID"] = t1
    sheets["07_gFID_raw"] = g

    # --- 02 Table 2: reconstruction (mean of per-cell) --------------------
    rmap, long_rows = {}, []
    for key, (fname, ckpt) in PERCELL.items():
        path = os.path.join(JP, fname)
        if not os.path.exists(path):
            continue
        w = pd.read_csv(path, index_col=0)
        for metric in w.index:
            for cell in w.columns:
                long_rows.append({"arm": key, "ckpt": ckpt, "metric": metric,
                                  "cell": cell, "value": w.loc[metric, cell]})
        inc = w.loc["ImgNet"].astype(float)
        for m in MODS:
            cs = [c for c in CELLS_BY_MOD[m] if c in inc.index]
            if cs:
                rmap[(key, m)] = sum(inc[c] for c in cs) / len(cs)
    sheets["06_Table2_rFID"] = paper_table(rmap)
    if long_rows:
        sheets["09_rFID_percell"] = pd.DataFrame(long_rows)

    # --- the rest pass through --------------------------------------------
    for sheet, fname in [("08_guidance", "results_pooled_guidance.csv"),
                         ("10_latent_by_modality", "results_pooled_latent_by_modality.csv")]:
        path = os.path.join(JP, fname)
        if os.path.exists(path):
            df = pd.read_csv(path)
            # A blank numeric cell cannot distinguish "not run yet" from "failed"
            # from "genuinely zero". The MIUA book solved this by typing "need job
            # submit" INTO the numeric column, which then could not be sorted or
            # charted. Keep the number column numeric and put the state beside it.
            if "fid_inception" in df.columns:
                df.insert(len(df.columns), "status",
                          df["fid_inception"].isna().map({True: "pending", False: "measured"}))
            sheets[sheet] = df

    # Slice tier lives in its own sheet, never merged with the per-cell one: all_T1
    # CONTAINS ukb_T1, so a single column holding both invites double counting.
    if os.path.exists(SLICE_SOURCE):
        sheets["01_recon_perslice"] = pd.read_csv(SLICE_SOURCE)
    # Generation gets its own pair, never merged with recon: they answer different
    # questions and share only the arm identity.
    if os.path.exists(GEN_CELL_SOURCE):
        sheets["02_gen_percell"] = pd.read_csv(GEN_CELL_SOURCE)
    if os.path.exists(GEN_SLICE_SOURCE):
        sheets["03_gen_perslice"] = pd.read_csv(GEN_SLICE_SOURCE)

    arms = arms_sheet()
    sheets["04_latent_stats"] = arms
    arms.to_csv(ARMS_OUT, index=False)

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    with pd.ExcelWriter(args.out, engine="openpyxl") as xl:
        for name in sorted(sheets):
            df = sheets[name]
            df.to_excel(xl, sheet_name=name[:31], index=False)
            ws = xl.sheets[name[:31]]
            ws.freeze_panes = "A2"
            for i, col in enumerate(df.columns, start=1):
                body = df[col].astype(str).str.len().max() if len(df) else 0
                width = min(max(len(str(col)), int(body) if body == body else 0) + 2, 70)
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
            # The README's prose columns need wrapping or they run off the screen.
            if name == "99_README":
                for row in ws.iter_rows(min_row=2):
                    for c in row:
                        c.alignment = Alignment(wrapText=True, vertical="top")

    if args.csv_dir:
        os.makedirs(args.csv_dir, exist_ok=True)
        for name in sorted(sheets):
            sheets[name].to_csv(os.path.join(args.csv_dir, f"{name}.csv"), index=False)
        print(f"wrote {len(sheets)} CSVs to {args.csv_dir}/")

    if args.core_out:
        core = ["00_recon_percell", "01_recon_perslice", "02_gen_percell",
                "03_gen_perslice", "04_latent_stats"]
        with pd.ExcelWriter(args.core_out, engine="openpyxl") as xl:
            for name in core:
                if name not in sheets:
                    continue
                df = sheets[name]
                df.to_excel(xl, sheet_name=name, index=False)
                ws = xl.sheets[name]
                ws.freeze_panes = "A2"
                for i, col in enumerate(df.columns, start=1):
                    body = df[col].astype(str).str.len().max() if len(df) else 0
                    width = min(max(len(str(col)), int(body) if body == body else 0) + 2, 40)
                    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        print(f"wrote {args.core_out}  ({', '.join(n for n in core if n in sheets)})")

    print(f"wrote {args.out}")
    for name in sorted(sheets):
        print(f"  {name:22s} {len(sheets[name]):4d} rows x {len(sheets[name].columns)} cols")


if __name__ == "__main__":
    main()
